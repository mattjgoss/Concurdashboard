import os
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def _default_template_path() -> str:
    """
    Resolve the Excel template path reliably in Azure/App Service.

    Expected structure:
      /home/site/wwwroot/
        main.py
        services/
          excel_export.py
        reports/
          accrual report.xlsx
    """
    here = os.path.dirname(os.path.abspath(__file__))
    return os.path.abspath(os.path.join(here, "..", "reports", "accrual report.xlsx"))


TEMPLATE_PATH = _default_template_path()


def _clear_data(ws: Worksheet, start_row: int = 2) -> None:
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def export_accruals_to_excel(
    unsubmitted_reports: List[Dict[str, Any]],
    unassigned_cards: List[Dict[str, Any]],
    card_totals_by_program: Optional[List[Dict[str, Any]]] = None,
    card_totals_by_user: Optional[List[Dict[str, Any]]] = None,
    meta: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    Populates the accrual report Excel template and returns XLSX bytes.
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(
            f"Excel template not found at '{TEMPLATE_PATH}'. "
            f"Ensure 'reports/accrual report.xlsx' is included in the zip at the correct path."
        )

    wb = load_workbook(TEMPLATE_PATH)

    # Optional meta sheet (only if template contains it)
    if meta and "Meta" in wb.sheetnames:
        ws_meta = wb["Meta"]
        ws_meta["A1"].value = "Generated"
        ws_meta["B1"].value = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%SZ")
        row = 3
        for k, v in meta.items():
            ws_meta.cell(row, 1).value = str(k)
            ws_meta.cell(row, 2).value = "" if v is None else str(v)
            row += 1

    # Sheet: "unsubnitted reports"
    ws_reports = wb["unsubnitted reports"]
    _clear_data(ws_reports)

    for row, r in enumerate(unsubmitted_reports, start=2):
        ws_reports.cell(row, 1).value = r.get("lastName")
        ws_reports.cell(row, 2).value = r.get("firstName")
        ws_reports.cell(row, 3).value = r.get("reportName")
        ws_reports.cell(row, 4).value = r.get("submitted")
        ws_reports.cell(row, 5).value = r.get("reportCreationDate")
        ws_reports.cell(row, 6).value = r.get("reportSubmissionDate")
        ws_reports.cell(row, 7).value = r.get("totalAmount")

    # Sheet: "unassigned card transactions"
    ws_cards = wb["unassigned card transactions"]
    _clear_data(ws_cards)

    for row, c in enumerate(unassigned_cards, start=2):
        ws_cards.cell(row, 1).value = c.get("cardProgramName") or c.get("cardProgramId")
        ws_cards.cell(row, 2).value = c.get("accountKey")
        ws_cards.cell(row, 3).value = c.get("lastFourDigits")
        ws_cards.cell(row, 4).value = c.get("transactionDate")
        ws_cards.cell(row, 5).value = c.get("postedDate")
        ws_cards.cell(row, 6).value = c.get("merchantName")
        ws_cards.cell(row, 7).value = c.get("description")
        ws_cards.cell(row, 8).value = c.get("postedAmount")
        ws_cards.cell(row, 9).value = c.get("postedCurrencyCode")

    # Optional totals sheet
    if card_totals_by_program or card_totals_by_user:
        ws_totals = wb["Card totals"] if "Card totals" in wb.sheetnames else wb.create_sheet("Card totals")
        ws_totals.delete_rows(1, ws_totals.max_row)

        row = 1
        ws_totals.cell(row, 1).value = "Totals by Program"
        row += 2

        ws_totals.cell(row, 1).value = "Program"
        ws_totals.cell(row, 2).value = "Count"
        ws_totals.cell(row, 3).value = "Total"
        ws_totals.cell(row, 4).value = "Currency"
        row += 1

        for p in card_totals_by_program or []:
            ws_totals.cell(row, 1).value = p.get("cardProgramName") or p.get("cardProgramId")
            ws_totals.cell(row, 2).value = p.get("count")
            ws_totals.cell(row, 3).value = p.get("total")
            ws_totals.cell(row, 4).value = p.get("currency")
            row += 1

        row += 2
        ws_totals.cell(row, 1).value = "Totals by User"
        row += 2

        ws_totals.cell(row, 1).value = "User"
        ws_totals.cell(row, 2).value = "Count"
        ws_totals.cell(row, 3).value = "Total"
        ws_totals.cell(row, 4).value = "Currency"
        row += 1

        for u in card_totals_by_user or []:
            ws_totals.cell(row, 1).value = u.get("userKey")
            ws_totals.cell(row, 2).value = u.get("count")
            ws_totals.cell(row, 3).value = u.get("total")
            ws_totals.cell(row, 4).value = u.get("currency")
            row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

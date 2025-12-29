from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from io import BytesIO
from typing import List, Dict, Any, Optional
from datetime import datetime

TEMPLATE_PATH = "/mnt/data/accrual report.xlsx"

def _clear_data(ws: Worksheet, start_row: int = 2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

def export_accruals_to_excel(
    unsubmitted_reports: List[Dict[str, Any]],
    unassigned_cards: List[Dict[str, Any]],
    card_totals_by_program: Optional[List[Dict[str, Any]]] = None,
    card_totals_by_user: Optional[List[Dict[str, Any]]] = None,
    meta: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    Populates the accrual report Excel template and returns XLSX bytes.
    """

    wb = load_workbook(TEMPLATE_PATH)

    # -------------------------------
    # Sheet: unsubnitted reports
    # -------------------------------
    ws_reports = wb["unsubnitted reports"]
    _clear_data(ws_reports)

    for row, r in enumerate(unsubmitted_reports, start=2):
        ws_reports.cell(row, 1).value = r.get("lastName")
        ws_reports.cell(row, 2).value = r.get("firstName")
        ws_reports.cell(row, 3).value = r.get("reportName")
        ws_reports.cell(row, 4).value = r.get("submitted")
        ws_reports.cell(row, 5).value = r.get("reportCreationDate")
        ws_reports.cell(row, 6).value = r.get("reportSubmissionDate")
        ws_reports.cell(row, 7).value = r.get("totalAmount")

    # -------------------------------
    # Sheet: Unassigned cards
    # -------------------------------
    ws_cards = wb["Unassigned cards"]
    _clear_data(ws_cards)

    for row, c in enumerate(unassigned_cards, start=2):
        ws_cards.cell(row, 1).value = c.get("cardProgramName")
        ws_cards.cell(row, 2).value = c.get("accountKey")  # lastSegment / fallback key
        ws_cards.cell(row, 3).value = c.get("lastFourDigits")
        ws_cards.cell(row, 4).value = c.get("postedAmount")

    # -------------------------------
    # Sheet: Card totals (optional)
    # -------------------------------
    if card_totals_by_program or card_totals_by_user:
        if "Card totals" in wb.sheetnames:
            ws_totals = wb["Card totals"]
            ws_totals.delete_rows(1, ws_totals.max_row)
        else:
            ws_totals = wb.create_sheet("Card totals")

        ws_totals["A1"] = "Card totals"
        ws_totals["A2"] = f"Generated: {datetime.now():%Y-%m-%d %H:%M}"

        if meta:
            ws_totals["A3"] = (
                f"Date range: {meta.get('dateFrom')} to {meta.get('dateTo')} "
                f"({meta.get('dateType')})"
            )

        # Totals by program
        row = 5
        ws_totals[row][0].value = "Totals by card program"
        row += 1
        ws_totals[row][0].value = "Card program"
        ws_totals[row][1].value = "Count"
        ws_totals[row][2].value = "Total amount"
        ws_totals[row][3].value = "Currency"
        row += 1

        for p in card_totals_by_program or []:
            ws_totals.cell(row, 1).value = p["cardProgramId"]
            ws_totals.cell(row, 2).value = p["count"]
            ws_totals.cell(row, 3).value = p["total"]
            ws_totals.cell(row, 4).value = p["currency"]
            row += 1

        # Totals by user
        row += 2
        ws_totals.cell(row, 1).value = "Totals by user (Employee ID else Account)"
        row += 1
        ws_totals.cell(row, 1).value = "User key"
        ws_totals.cell(row, 2).value = "Count"
        ws_totals.cell(row, 3).value = "Total amount"
        ws_totals.cell(row, 4).value = "Currency"
        row += 1

        for u in card_totals_by_user or []:
            ws_totals.cell(row, 1).value = u["userKey"]
            ws_totals.cell(row, 2).value = u["count"]
            ws_totals.cell(row, 3).value = u["total"]
            ws_totals.cell(row, 4).value = u["currency"]
            row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()

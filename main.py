print("#### LOADED MAIN FROM:", __file__)

import os
import time
from collections import defaultdict
from datetime import datetime, date
from io import BytesIO
from typing import Optional, List, Dict, Any

import requests
from dateutil.parser import isoparse
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel

# IMPORTANT: requires services/__init__.py and services/identity_service.py in the zip
from services.identity_service import get_secret, keyvault_status

BUILD_FINGERPRINT = os.getenv("SCM_COMMIT_ID") or os.getenv("WEBSITE_DEPLOYMENT_ID") or "unknown"
print("RUN_FROM_PACKAGE =", os.getenv("WEBSITE_RUN_FROM_PACKAGE"))
print("PWD =", os.getcwd())

# ======================================================
# FASTAPI APP
# ======================================================

app = FastAPI(title="Concur Accruals API")

# ======================================================
# LOCAL FILES
# ======================================================

TEMPLATE_PATH = "reports/accrual report.xlsx"

# ======================================================
# HELPERS
# ======================================================

def safe_body(resp: requests.Response, limit: int = 1000) -> str:
    """Return a safe, truncated response body for diagnostics."""
    try:
        txt = resp.text or ""
        return txt[:limit]
    except Exception:
        return "<unreadable>"

# ======================================================
# KEY VAULT ACCESS (thin wrapper)
# ======================================================

def kv(name: str, fallback: Optional[str] = None) -> Optional[str]:
    """
    Fetch a secret from Key Vault (via services.identity_service.get_secret).
    If Key Vault read fails or secret is missing/blank, return fallback.
    """
    try:
        val = get_secret(name)
        if val is None:
            return fallback
        if isinstance(val, str) and val.strip() == "":
            return fallback
        return val
    except Exception:
        return fallback

def concur_base_url() -> str:
    # Example secret value: https://us2.api.concursolutions.com
    base = kv("concur-api-base-url", os.getenv("CONCUR_API_BASE_URL"))
    if not base:
        raise HTTPException(status_code=500, detail={
            "status": "error",
            "error": "missing_config",
            "missing": ["concur-api-base-url / CONCUR_API_BASE_URL"],
        })
    return str(base).rstrip("/")

def concur_token_url() -> str:
    # Example secret value: https://us2.api.concursolutions.com/oauth2/v0/token
    # If not set, derive from base URL.
    token = kv("concur-token-url", os.getenv("CONCUR_TOKEN_URL"))
    if token:
        return str(token).rstrip("/")
    return f"{concur_base_url()}/oauth2/v0/token"

# ======================================================
# CONCUR OAUTH (REFRESH TOKEN FLOW)
# ======================================================

class ConcurOAuthClient:
    def __init__(self):
        self.access_token: Optional[str] = None
        self.expires_at: float = 0.0

    def get_access_token(self) -> str:
        now = time.time()
        if self.access_token and now < self.expires_at - 60:
            return self.access_token

        token_url = concur_token_url()

        client_id = kv("concur-client-id", os.getenv("CONCUR_CLIENT_ID"))
        client_secret = kv("concur-client-secret", os.getenv("CONCUR_CLIENT_SECRET"))
        refresh_token = kv("concur-refresh-token", os.getenv("CONCUR_REFRESH_TOKEN"))

        missing = [k for k, v in {
            "concur-client-id / CONCUR_CLIENT_ID": client_id,
            "concur-client-secret / CONCUR_CLIENT_SECRET": client_secret,
            "concur-refresh-token / CONCUR_REFRESH_TOKEN": refresh_token,
        }.items() if not v]
        if missing:
            raise HTTPException(status_code=500, detail={
                "status": "error",
                "error": "missing_config",
                "missing": missing,
            })

        # Concur refresh-token flow: client_id/client_secret can be sent either in
        # Authorization header (basic auth) or in form body depending on app config.
        # Your earlier pattern used form fields; we keep it (works for many Concur apps).
        resp = requests.post(
            token_url,
            data={
                "grant_type": "refresh_token",
                "refresh_token": refresh_token,
                "client_id": client_id,
                "client_secret": client_secret,
            },
            timeout=20,
        )

        if resp.status_code in (400, 401, 403):
            raise HTTPException(
                status_code=502,
                detail={
                    "stage": "token",
                    "error": "concur_oauth_failed",
                    "status_code": resp.status_code,
                    "token_endpoint": token_url,
                    "body": safe_body(resp),
                    "hint": "Check Key Vault secrets + Concur app refresh token + client credentials.",
                },
            )

        resp.raise_for_status()
        data = resp.json()

        if "access_token" not in data:
            raise HTTPException(
                status_code=502,
                detail={
                    "stage": "token",
                    "error": "no_access_token_in_response",
                    "token_endpoint": token_url,
                    "token_response_keys": list(data.keys()),
                    "body": safe_body(resp),
                },
            )

        self.access_token = data["access_token"]
        self.expires_at = now + float(data.get("expires_in", 1800))
        return self.access_token

oauth = ConcurOAuthClient()

def concur_headers() -> Dict[str, str]:
    return {"Authorization": f"Bearer {oauth.get_access_token()}", "Accept": "application/json"}

# ======================================================
# REQUEST MODELS
# ======================================================

class AccrualsSearchRequest(BaseModel):
    orgUnit1: Optional[str] = None
    orgUnit2: Optional[str] = None
    orgUnit3: Optional[str] = None
    orgUnit4: Optional[str] = None
    orgUnit5: Optional[str] = None
    orgUnit6: Optional[str] = None
    custom21: Optional[str] = None

class CardTotalsRequest(AccrualsSearchRequest):
    transactionDateFrom: str
    transactionDateTo: str
    dateType: str  # TRANSACTION | POSTED | BILLING

# ======================================================
# IDENTITY v4.1 (USER RESOLUTION)
# ======================================================

def build_identity_filter(req: Any) -> Optional[str]:
    filters = []

    for i in range(1, 7):
        val = getattr(req, f"orgUnit{i}", None)
        if val:
            filters.append(
                f'urn:ietf:params:scim:schemas:extension:spend:2.0:User:orgUnit{i} eq "{val}"'
            )

    if getattr(req, "custom21", None):
        filters.append(
            'urn:ietf:params:scim:schemas:extension:spend:2.0:User:customData'
            f'[id eq "custom21" and value eq "{req.custom21}"]'
        )

    return " and ".join(filters) if filters else None

def get_users(filter_expression: Optional[str]) -> List[Dict]:
    url = f"{concur_base_url()}/profile/identity/v4.1/Users"
    params = {
        "attributes": (
            "id,displayName,userName,emails.value,"
            "urn:ietf:params:scim:schemas:extension:enterprise:2.0:User"
        )
    }
    if filter_expression:
        params["filter"] = filter_expression

    resp = requests.get(url, headers=concur_headers(), params=params, timeout=30)
    resp.raise_for_status()
    return resp.json().get("Resources", []) or []

# ======================================================
# EXPENSE REPORTS v4
# ======================================================

def get_expense_reports(user_id: str) -> List[Dict]:
    url = f"{concur_base_url()}/expensereports/v4/users/{user_id}/reports"
    resp = requests.get(url, headers=concur_headers(), timeout=30)
    resp.raise_for_status()
    return resp.json().get("Items", []) or []

def filter_unsubmitted_reports(reports: List[Dict]) -> List[Dict]:
    results = []
    for r in reports:
        payment_status_id = r.get("paymentStatusId")

        # Exclude reports already paid or processing
        if payment_status_id in ("P_PAID", "P_PROC"):
            continue

        results.append(
            {
                "lastName": r.get("owner", {}).get("lastName"),
                "firstName": r.get("owner", {}).get("firstName"),
                "reportName": r.get("name"),
                "submitted": r.get("approvalStatus") == "Submitted",
                "reportCreationDate": r.get("creationDate"),
                "reportSubmissionDate": r.get("submitDate"),
                "paymentStatusId": payment_status_id,
                "totalAmount": (r.get("totalAmount") or {}).get("value"),
            }
        )
    return results

# ======================================================
# CARDS v4
# ======================================================

def get_card_transactions(user_id: str, date_from: str, date_to: str) -> List[Dict]:
    url = f"{concur_base_url()}/cards/v4/users/{user_id}/transactions"
    params = {"transactionDateFrom": date_from, "transactionDateTo": date_to}
    resp = requests.get(url, headers=concur_headers(), params=params, timeout=30)
    resp.raise_for_status()
    return resp.json().get("Items", []) or []

def filter_unassigned_cards(transactions: List[Dict]) -> List[Dict]:
    results = []
    for t in transactions:
        if t.get("expenseId") or t.get("reportId"):
            continue

        account = t.get("account") or {}
        payment_type = account.get("paymentType") or {}

        posted = t.get("postedAmount") or {}
        results.append(
            {
                "cardProgramId": payment_type.get("id"),
                "accountKey": account.get("lastSegment"),
                "lastFourDigits": account.get("lastFourDigits"),
                "postedAmount": posted.get("value"),
                "currencyCode": posted.get("currencyCode"),
            }
        )
    return results

# ======================================================
# CARD TOTALS LOGIC
# ======================================================

def extract_date(txn: Dict, date_type: str) -> date:
    if date_type == "POSTED":
        return isoparse(txn["postedDate"]).date()
    if date_type == "BILLING":
        return isoparse(txn["statement"]["billingDate"]).date()
    return isoparse(txn["transactionDate"]).date()

def compute_card_totals(transactions: List[Dict], from_date: date, to_date: date, date_type: str):
    by_program = defaultdict(lambda: {"count": 0, "total": 0.0, "currency": ""})
    by_user = defaultdict(lambda: {"count": 0, "total": 0.0, "currency": ""})

    for t in transactions:
        if date_type == "BILLING":
            stmt = t.get("statement") or {}
            if not stmt.get("billingDate"):
                continue
        if date_type == "POSTED" and not t.get("postedDate"):
            continue
        if date_type == "TRANSACTION" and not t.get("transactionDate"):
            continue

        d = extract_date(t, date_type)
        if not (from_date <= d <= to_date):
            continue

        posted_amount = t.get("postedAmount") or {}
        amount = float(posted_amount.get("value") or 0.0)
        currency = str(posted_amount.get("currencyCode") or "")

        account = t.get("account") or {}
        payment_type = account.get("paymentType") or {}
        program = str(payment_type.get("id") or "")

        employee_id = t.get("employeeId")
        user_key = employee_id if employee_id else f'{account.get("lastSegment")} ({program})'

        by_program[program]["count"] += 1
        by_program[program]["total"] += amount
        by_program[program]["currency"] = currency

        by_user[user_key]["count"] += 1
        by_user[user_key]["total"] += amount
        by_user[user_key]["currency"] = currency

    return {
        "totalsByProgram": [{"cardProgramId": k, **v} for k, v in by_program.items()],
        "totalsByUser": [{"userKey": k, **v} for k, v in by_user.items()],
    }

# ======================================================
# EXCEL EXPORT
# ======================================================

def export_card_totals_excel(totals: Dict, meta: Dict) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)

    ws = wb["Card totals"] if "Card totals" in wb.sheetnames else wb.create_sheet("Card totals")
    ws.delete_rows(1, ws.max_row)

    ws["A1"] = "Card totals"
    ws["A2"] = f"Generated {datetime.now():%Y-%m-%d %H:%M}"
    ws["A3"] = f'Date range: {meta["from"]} to {meta["to"]} ({meta["type"]})'

    ws.append([])
    ws.append(["Card program", "Count", "Total", "Currency"])

    for p in totals["totalsByProgram"]:
        ws.append([p.get("cardProgramId"), p.get("count"), p.get("total"), p.get("currency")])

    ws.append([])
    ws.append(["User key", "Count", "Total", "Currency"])

    for u in totals["totalsByUser"]:
        ws.append([u.get("userKey"), u.get("count"), u.get("total"), u.get("currency")])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()

# ======================================================
# API ENDPOINTS
# ======================================================

@app.post("/api/accruals/search")
def accruals_search(req: AccrualsSearchRequest):
    filter_expr = build_identity_filter(req)
    users = get_users(filter_expr)

    unsubmitted_reports = []
    unassigned_cards = []

    for u in users:
        reports = get_expense_reports(u["id"])
        unsubmitted_reports.extend(filter_unsubmitted_reports(reports))

        txns = get_card_transactions(u["id"], "2000-01-01", date.today().isoformat())
        unassigned_cards.extend(filter_unassigned_cards(txns))

    return {
        "summary": {
            "unsubmittedReportCount": len(unsubmitted_reports),
            "unassignedCardCount": len(unassigned_cards),
        },
        "unsubmittedReports": unsubmitted_reports,
        "unassignedCards": unassigned_cards,
    }

@app.post("/api/cardtotals/export")
def card_totals_export(req: CardTotalsRequest):
    date_type = (req.dateType or "").upper()
    if date_type not in ("TRANSACTION", "POSTED", "BILLING"):
        raise HTTPException(status_code=400, detail="dateType must be TRANSACTION, POSTED, or BILLING")

    filter_expr = build_identity_filter(req)
    users = get_users(filter_expr)

    all_txns = []
    for u in users:
        all_txns.extend(get_card_transactions(u["id"], req.transactionDateFrom, req.transactionDateTo))

    totals = compute_card_totals(
        all_txns,
        date.fromisoformat(req.transactionDateFrom),
        date.fromisoformat(req.transactionDateTo),
        date_type,
    )

    xlsx = export_card_totals_excel(
        totals,
        {"from": req.transactionDateFrom, "to": req.transactionDateTo, "type": date_type},
    )

    filename = f"Concur_Card_Totals_{datetime.now():%Y%m%d_%H%M}.xlsx"

    return StreamingResponse(
        BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@app.get("/kv-test")
def kv_test():
    """
    Simple Key Vault read test.
    Returns:
      - status: ok (if read works)
      - status: error (if KV config/MI/permissions fail)
    """
    try:
        client_id = kv("concur-client-id")
        return {"status": "ok", "client_id_exists": bool(client_id), "keyvault": keyvault_status()}
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail={"message": f"Key Vault read failed: {str(e)}", "keyvault": keyvault_status()},
        )

@app.get("/build")
def build():
    return {
        "fingerprint": BUILD_FINGERPRINT,
        "file": __file__,
        "cwd": os.getcwd(),
        "run_from_package": os.getenv("WEBSITE_RUN_FROM_PACKAGE"),
        "scm_build": os.getenv("SCM_DO_BUILD_DURING_DEPLOYMENT"),
        "scm_commit": os.getenv("SCM_COMMIT_ID"),
        "deployment_id": os.getenv("WEBSITE_DEPLOYMENT_ID"),
    }

@app.get("/api/concur/auth-test")
def concur_auth_test():
    # 1) Load config/secrets (Key Vault first, env fallback)
    base_url = kv("concur-api-base-url", os.getenv("CONCUR_API_BASE_URL"))
    if not base_url:
        raise HTTPException(status_code=500, detail={
            "status": "error",
            "error": "missing_config",
            "missing": ["concur-api-base-url / CONCUR_API_BASE_URL"],
        })

    token_url = kv("concur-token-url", os.getenv("CONCUR_TOKEN_URL")) or f"{str(base_url).rstrip('/')}/oauth2/v0/token"

    client_id = kv("concur-client-id", os.getenv("CONCUR_CLIENT_ID"))
    client_secret = kv("concur-client-secret", os.getenv("CONCUR_CLIENT_SECRET"))
    refresh_token = kv("concur-refresh-token", os.getenv("CONCUR_REFRESH_TOKEN"))

    missing = [k for k, v in {
        "client_id": client_id,
        "client_secret": client_secret,
        "refresh_token": refresh_token,
    }.items() if not v]
    if missing:
        raise HTTPException(status_code=500, detail={
            "status": "error",
            "error": "missing_config",
            "missing": missing,
        })

    # 2) Refresh token -> access token
    data = {"grant_type": "refresh_token", "refresh_token": refresh_token}

    r = requests.post(token_url, data=data, auth=(client_id, client_secret), timeout=30)
    if r.status_code != 200:
        raise HTTPException(status_code=502, detail={
            "stage": "token",
            "status_code": r.status_code,
            "body": safe_body(r),
            "token_endpoint": token_url,
        })

    tok = r.json()
    access_token = tok.get("access_token")
    if not access_token:
        raise HTTPException(status_code=502, detail={
            "stage": "token",
            "error": "No access_token in response",
            "token_response_keys": list(tok.keys()),
        })

    # 3) Call a lightweight Concur API endpoint (Identity)
    url = f"{str(base_url).rstrip('/')}/profile/identity/v4.1/Users?startIndex=1&count=1"
    h = {"Authorization": f"Bearer {access_token}", "Accept": "application/json"}
    u = requests.get(url, headers=h, timeout=30)

    return {
        "status": "ok" if u.status_code == 200 else "fail",
        "token_endpoint": token_url,
        "identity_test_url": url,
        "identity_status_code": u.status_code,
        "identity_snippet": (u.text or "")[:300],
        "expires_in": tok.get("expires_in"),
        "token_type": tok.get("token_type"),
        "scope": tok.get("scope"),
    }

@app.get("/debug/routes")
def debug_routes():
    return sorted([getattr(r, "path", "") for r in app.router.routes])

@app.get("/debug/deploy")
def debug_deploy():
    return {"deploy": "auth-test-v1", "utc": datetime.utcnow().isoformat() + "Z"}
